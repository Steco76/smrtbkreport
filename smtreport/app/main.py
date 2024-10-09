# app/main.py

import streamlit as st
import pandas as pd
import openpyxl
import io

def load_excel_file(uploaded_file):
    try:
        # Lade die Excel-Datei mit openpyxl
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        
        # Bestimme das relevante Tabellenblatt
        if "Tantiemen insgesamt" in wb.sheetnames:
            sheet_name = "Tantiemen insgesamt"
        elif "Gesamteinnahmen" in wb.sheetnames:
            sheet_name = "Gesamteinnahmen"
        else:
            st.error(f"Kein passendes Tabellenblatt in der Datei {uploaded_file.name} gefunden.")
            return None
        
        sheet = wb[sheet_name]
        
        # Extrahiere den Verkaufszeitraum aus Zelle B1
        sales_period = sheet['B1'].value
        
        # Lade die Daten ab Zeile 2 (header=1)
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=1)
        
        # Spaltenüberschriften bereinigen (Leerzeichen entfernen)
        df.columns = df.columns.str.strip()
        
        # Überprüfe und benenne die Einnahmenspalte um
        if "Einnahmen" in df.columns:
            df.rename(columns={"Einnahmen": "Tantiemen"}, inplace=True)
        elif "Tantiemen" in df.columns:
            df.rename(columns={"Tantiemen": "Tantiemen"}, inplace=True)
        else:
            st.error(f"Die Datei {uploaded_file.name} enthält keine Spalte 'Einnahmen' oder 'Tantiemen'.")
            return None
        
        # Überprüfe, ob die DataFrame leer ist (keine Datenzeilen)
        if df.empty:
            st.warning(f"Die Datei {uploaded_file.name} enthält keine Datenzeilen.")
            return None
        
        # Füge die neue Spalte 'Verkaufszeitraum' hinzu
        df['Verkaufszeitraum'] = sales_period
        
        # Konvertiere 'Verkaufszeitraum' von Text zu Datum und füge zusätzliche Spalten hinzu
        df = convert_sales_period_to_date(df)
        
        # Füge zusätzliche Spalten basierend auf 'Zahlungsplan' hinzu
        df = add_additional_columns(df)
        
        # Erstelle die 'Gesamtverkäufe' Spalte und füge sie zwischen 'Tantiemen' und 'E-Books' ein
        df['Gesamtverkäufe'] = df['E-Books'] + df['Paperback/Hardcover']
        cols = list(df.columns)
        tantiemen_index = cols.index('Tantiemen')
        cols.insert(tantiemen_index + 1, cols.pop(cols.index('Gesamtverkäufe')))
        df = df[cols]
        
        return df
    except Exception as e:
        st.error(f"Fehler beim Laden der Datei {uploaded_file.name}: {e}")
        return None

def convert_sales_period_to_date(df):
    # Mapping der deutschen Monatsnamen zu Monatsnummern
    month_mapping = {
        'Januar': '01',
        'Februar': '02',
        'März': '03',
        'April': '04',
        'Mai': '05',
        'Juni': '06',
        'Juli': '07',
        'August': '08',
        'September': '09',
        'Oktober': '10',
        'November': '11',
        'Dezember': '12'
    }
    
    # Mapping der Monatsnummern zu deutschen Monatsnamen
    month_num_to_de = {
        1: 'Januar',
        2: 'Februar',
        3: 'März',
        4: 'April',
        5: 'Mai',
        6: 'Juni',
        7: 'Juli',
        8: 'August',
        9: 'September',
        10: 'Oktober',
        11: 'November',
        12: 'Dezember'
    }
    
    def parse_month_year_de(text):
        """
        Wandelt den Verkaufszeitraum im Format 'Monat Jahr' (z.B. 'Januar 2024') in ein Datum um.
        Das Datum wird auf den ersten Tag des Monats gesetzt.
        """
        try:
            parts = text.split()  # Erwartet ['Monat', 'Jahr']
            if len(parts) != 2:
                return pd.NaT
            month_str, year_str = parts
            month_num = month_mapping.get(month_str.capitalize(), None)
            if month_num is None:
                return pd.NaT
            return pd.Timestamp(f"{year_str}-{month_num}-01")
        except:
            return pd.NaT
    
    # Wende die Parsing-Funktion auf die 'Verkaufszeitraum'-Spalte an
    df['Verkaufszeitraum'] = df['Verkaufszeitraum'].apply(parse_month_year_de)
    
    # Füge die Spalten 'Monat', 'Jahr' und 'Monat_num' hinzu
    df['Monat'] = df['Verkaufszeitraum'].dt.month.map(month_num_to_de)
    df['Jahr'] = df['Verkaufszeitraum'].dt.year
    df['Monat_num'] = df['Verkaufszeitraum'].dt.month
    
    # Drope Zeilen mit fehlendem 'Jahr'
    missing_jahr = df['Jahr'].isna().sum()
    if missing_jahr > 0:
        st.warning(f"{missing_jahr} Zeile(n) haben einen ungültigen Verkaufszeitraum und werden ignoriert.")
        df = df.dropna(subset=['Jahr'])
    
    # Sicherstellen, dass 'Jahr' integer ist
    if not df.empty:
        try:
            df['Jahr'] = df['Jahr'].astype(int)
        except Exception as e:
            st.error(f"Fehler bei der Umwandlung von 'Jahr' in Integer: {e}")
            return df  # Rückgabe ohne Konvertierung
    
    return df

def add_additional_columns(df):
    """
    Fügt die Spalten 'E-Books', 'Paperback/Hardcover', 'Gelesene Seiten' und 'Bonus' basierend auf 'Zahlungsplan' hinzu.
    """
    # Initialisiere die neuen Spalten mit 0
    df['E-Books'] = 0
    df['Paperback/Hardcover'] = 0
    df['Gelesene Seiten'] = 0
    df['Bonus'] = 0  # Neue Spalte 'Bonus' initialisiert mit 0
    
    # Bedingungen
    df.loc[df['Zahlungsplan'] == "Standard", 'E-Books'] = df['Netto verkaufte Einheiten oder gelesene KENP-Seiten**']
    df.loc[df['Zahlungsplan'].isin(["Standard – Taschenbuch", "Standard – Gebundene Ausgabe"]), 'Paperback/Hardcover'] = df['Netto verkaufte Einheiten oder gelesene KENP-Seiten**']
    df.loc[df['Zahlungsplan'] == "Gelesene KENP-Seiten (Kindle Edition Normalized Pages Read)", 'Gelesene Seiten'] = df['Netto verkaufte Einheiten oder gelesene KENP-Seiten**']
    
    # Bedingung für 'Bonus' Spalte
    df.loc[df['Zahlungsplan'].isin(["All-Stars-Bonus", "All Star Bonus"]), 'Bonus'] = df['Tantiemen']
    
    return df

def aggregate_einnahmen_pro_autor_wahrung(df):
    """
    Aggregiert die Gesamtsumme der Einnahmen, Gesamtverkäufe, E-Books, Paperback/Hardcover, Gelesene Seiten und Bonus
    pro Autor, Währung, Jahr, Monat und Titel.
    """
    aggregated_df = df.groupby(['Autor', 'Währung', 'Jahr', 'Monat', 'Monat_num', 'Titel'])[
        ['Tantiemen', 'Gesamtverkäufe', 'E-Books', 'Paperback/Hardcover', 'Gelesene Seiten', 'Bonus']
    ].sum().reset_index()
    return aggregated_df

def format_eu_number(x, decimal_places=0):
    """
    Formatiert eine Zahl im EU-Format.
    - decimal_places: Anzahl der Dezimalstellen.
    """
    if pd.isna(x):
        return ''
    try:
        if decimal_places > 0:
            formatted = f"{x:,.{decimal_places}f}".replace(",", " ").replace(".", ",").replace(" ", ".")
        else:
            formatted = f"{x:,}".replace(",", ".")
        return formatted
    except:
        return str(x)

def main():
    
    # Überschrift und Beschreibung (optional)
    # st.title("📚 Übersicht Buchverkäufe")
    # st.write("Laden Sie mehrere Excel-Dateien hoch und verarbeiten Sie die Daten.")
    
    # Mapping der deutschen Monatsnamen zu Monatsnummern für Sortierung
    month_order = {
        'Januar': 1,
        'Februar': 2,
        'März': 3,
        'April': 4,
        'Mai': 5,
        'Juni': 6,
        'Juli': 7,
        'August': 8,
        'September': 9,
        'Oktober': 10,
        'November': 11,
        'Dezember': 12
    }
    
    # Initialisiere den Session State für den File Uploader Key, falls nicht vorhanden
    if 'file_uploader_key' not in st.session_state:
        st.session_state.file_uploader_key = 0
    
    # Datei-Upload erlauben mit dynamischem Key
    uploaded_files = st.file_uploader(
        "📂 Excel-Datei(en) auswählen:",
        type=["xlsx"],
        accept_multiple_files=True,
        # help="Es können mehrere Dateien gleichzeitg ausgewählt werden.",
        key=f"uploaded_files_{st.session_state.file_uploader_key}"
    )
    
    # Filtern von doppelten Dateien innerhalb der aktuellen Upload
    if uploaded_files:
        unique_uploaded_files = []
        duplicate_files = []
        seen_filenames = set()
        for file in uploaded_files:
            if file.name in seen_filenames:
                duplicate_files.append(file.name)
            else:
                unique_uploaded_files.append(file)
                seen_filenames.add(file.name)
        
        # Informiere den Benutzer über doppelte Dateien
        if duplicate_files:
            st.warning(f"Doppelte Dateien wurden entfernt: {', '.join(duplicate_files)}")
    else:
        unique_uploaded_files = []
    
    # Zwei Buttons nebeneinander
    col1, col2 = st.columns([1, 1])
    
    with col1:
        if st.button("✅ Daten bearbeiten"):
            if not unique_uploaded_files:
                st.error("Bitte laden Sie mindestens eine Excel-Datei hoch.")
            else:
                combined_data = []
                for uploaded_file in unique_uploaded_files:
                    df = load_excel_file(uploaded_file)
                    if df is not None:
                        combined_data.append(df)
                
                if combined_data:
                    # Kombiniere alle DataFrames
                    combined_df = pd.concat(combined_data, ignore_index=True)
                    
                    # **Neuer Abschnitt für die Zusammenfassung**
                    st.success(f"{len(combined_data)} Datei(en) geladen mit insgesamt {len(combined_df)} Datensätzen.")
                    # **Ende des neuen Abschnitts**
                    
                    # Aggregation: Gesamtsumme der Einnahmen, Gesamtverkäufe, E-Books, Paperback/Hardcover, Gelesene Seiten und Bonus pro Autor, Währung, Jahr, Monat und Titel
                    aggregated_df = aggregate_einnahmen_pro_autor_wahrung(combined_df)
                    
                    # Speichern der aggregierten Daten in Session State für spätere Verwendung
                    st.session_state['aggregated_einnahmen'] = aggregated_df
                    
                    # Reset des File Uploader durch Aktualisieren des Keys
                    st.session_state.file_uploader_key += 1
                    
                else:
                    st.error("Keine gültigen Daten gefunden oder Fehler beim Verarbeiten der Dateien.")
    
    # Zugriff auf den aggregierten DataFrame
    aggregated_df = st.session_state.get('aggregated_einnahmen', pd.DataFrame())
    
    if not aggregated_df.empty:
        
        # Auswahl von Autor, Titel, Jahr, Monat, Währung und Bonus zur Anzeige der Metriken
        
        # Autor Auswahl
        autor_unique = sorted(aggregated_df['Autor'].unique())
        if len(autor_unique) > 1:
            autor_options = ["Alle"] + autor_unique
            autor_default = "Alle"
        else:
            autor_options = autor_unique
            autor_default = autor_unique[0]
        
        autor = st.selectbox("🔍 Wähle einen Autor", autor_options, index=autor_options.index(autor_default))
        
        # Auswahl von Titel, abhängig von Autor
        if autor == "Alle":
            titel_unique = sorted(aggregated_df['Titel'].unique())
        else:
            titel_unique = sorted(aggregated_df[aggregated_df['Autor'] == autor]['Titel'].unique())
        
        if len(titel_unique) > 1:
            titel_options = ["Alle"] + titel_unique
            titel_default = "Alle"
        elif len(titel_unique) == 1:
            titel_options = titel_unique
            titel_default = titel_unique[0]
        else:
            titel_options = ["Keine Titel verfügbar"]
            titel_default = "Keine Titel verfügbar"
        
        titel = st.selectbox("📖 Wähle einen Titel", titel_options, index=0)
        
        # Auswahl von Jahr, abhängig von Autor und Titel
        if autor == "Alle" and titel == "Alle":
            jahre_unique = sorted(aggregated_df['Jahr'].unique())
        elif autor == "Alle":
            jahre_unique = sorted(aggregated_df[aggregated_df['Titel'] == titel]['Jahr'].unique())
        elif titel == "Alle":
            jahre_unique = sorted(aggregated_df[aggregated_df['Autor'] == autor]['Jahr'].unique())
        else:
            jahre_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Titel'] == titel)]['Jahr'].unique())
        
        if len(jahre_unique) > 1:
            jahr_options = ["Alle"] + [str(jahr) for jahr in jahre_unique]
            jahr_default = "Alle"
        elif len(jahre_unique) == 1:
            jahr_options = [str(jahre_unique[0])]
            jahr_default = jahr_options[0]
        else:
            jahr_options = ["Keine Jahre verfügbar"]
            jahr_default = "Keine Jahre verfügbar"
        
        jahr = st.selectbox("📆 Wähle ein Jahr", jahr_options, index=0)
        
        # Auswahl von Monat, abhängig von Autor, Titel und Jahr
        if autor == "Alle" and titel == "Alle" and jahr == "Alle":
            monate_unique = sorted(aggregated_df['Monat'].unique(), key=lambda x: month_order.get(x, 13))
        elif autor == "Alle" and titel == "Alle":
            if jahr != "Alle":
                monate_unique = sorted(aggregated_df[aggregated_df['Jahr'] == int(jahr)]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
            else:
                monate_unique = sorted(aggregated_df['Monat'].unique(), key=lambda x: month_order.get(x, 13))
        elif autor == "Alle":
            if titel != "Alle" and jahr != "Alle":
                monate_unique = sorted(aggregated_df[(aggregated_df['Titel'] == titel) & (aggregated_df['Jahr'] == int(jahr))]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
            elif titel != "Alle":
                monate_unique = sorted(aggregated_df[aggregated_df['Titel'] == titel]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
            elif jahr != "Alle":
                monate_unique = sorted(aggregated_df[aggregated_df['Jahr'] == int(jahr)]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
            else:
                monate_unique = sorted(aggregated_df['Monat'].unique(), key=lambda x: month_order.get(x, 13))
        elif titel == "Alle":
            if jahr != "Alle":
                monate_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Jahr'] == int(jahr))]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
            else:
                monate_unique = sorted(aggregated_df[aggregated_df['Autor'] == autor]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
        else:
            if jahr != "Alle":
                monate_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Titel'] == titel) & (aggregated_df['Jahr'] == int(jahr))]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
            else:
                monate_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Titel'] == titel)]['Monat'].unique(), key=lambda x: month_order.get(x, 13))
        
        if len(monate_unique) > 1:
            monat_options = ["Alle"] + monate_unique
            monat_default = "Alle"
        elif len(monate_unique) == 1:
            monat_options = [monate_unique[0]]
            monat_default = monat_options[0]
        else:
            monat_options = ["Keine Monate verfügbar"]
            monat_default = "Keine Monate verfügbar"
        
        monat = st.selectbox("🗓️ Wähle einen Monat", monat_options, index=0)
        
        # Auswahl von Währung, abhängig von allen vorherigen Selektierungen
        if autor == "Alle" and titel == "Alle" and jahr == "Alle" and monat == "Alle":
            währung_unique = sorted(aggregated_df['Währung'].unique())
        elif autor == "Alle" and titel == "Alle" and jahr == "Alle":
            währung_unique = sorted(aggregated_df[aggregated_df['Monat'] == monat]['Währung'].unique()) if monat != "Alle" else sorted(aggregated_df['Währung'].unique())
        elif autor == "Alle" and titel == "Alle":
            if jahr != "Alle":
                währung_unique = sorted(aggregated_df[(aggregated_df['Jahr'] == int(jahr)) & (aggregated_df['Monat'] == monat)]['Währung'].unique()) if monat != "Alle" else sorted(aggregated_df[aggregated_df['Jahr'] == int(jahr)]['Währung'].unique())
            else:
                währung_unique = sorted(aggregated_df['Währung'].unique())
        elif autor == "Alle":
            if titel != "Alle" and jahr != "Alle":
                währung_unique = sorted(aggregated_df[(aggregated_df['Titel'] == titel) & (aggregated_df['Jahr'] == int(jahr)) & (aggregated_df['Monat'] == monat)]['Währung'].unique()) if monat != "Alle" else sorted(aggregated_df[(aggregated_df['Titel'] == titel) & (aggregated_df['Jahr'] == int(jahr))]['Währung'].unique())
            elif titel != "Alle":
                währung_unique = sorted(aggregated_df[(aggregated_df['Titel'] == titel)]['Währung'].unique())
            elif jahr != "Alle":
                währung_unique = sorted(aggregated_df[(aggregated_df['Jahr'] == int(jahr))]['Währung'].unique())
            else:
                währung_unique = sorted(aggregated_df['Währung'].unique())
        elif titel == "Alle":
            if jahr != "Alle":
                währung_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Jahr'] == int(jahr)) & (aggregated_df['Monat'] == monat)]['Währung'].unique()) if monat != "Alle" else sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Jahr'] == int(jahr))]['Währung'].unique())
            else:
                währung_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor)]['Währung'].unique())
        else:
            if jahr != "Alle":
                währung_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Titel'] == titel) & (aggregated_df['Jahr'] == int(jahr)) & (aggregated_df['Monat'] == monat)]['Währung'].unique()) if monat != "Alle" else sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Titel'] == titel) & (aggregated_df['Jahr'] == int(jahr))]['Währung'].unique())
            else:
                währung_unique = sorted(aggregated_df[(aggregated_df['Autor'] == autor) & (aggregated_df['Titel'] == titel)]['Währung'].unique())

        # Hinzufügen von "Alle" zurück und Setzen des Defaultwerts auf "EUR"
        if len(währung_unique) > 1:
            währung_options = ["Alle"] + währung_unique
            if "EUR" in währung_unique:
                default_index = währung_options.index("EUR")
            else:
                default_index = 0  # Standard auf "Alle" falls "EUR" nicht vorhanden
        elif len(währung_unique) == 1:
            währung_options = [währung_unique[0]]
            default_index = 0
        else:
            währung_options = ["Keine Währung verfügbar"]
            default_index = 0

        währung = st.selectbox(
            "💱 Wähle eine Währung",
            währung_options,
            index=default_index
        )
        
        # 3. Bonus Filter hinzufügen
        bonus_filter = st.selectbox(
            "🎁 Bonus Filter",
            ["Alle", "Mit Bonus", "Ohne Bonus"],
            index=0
        )
        
        # Filtere die Daten basierend auf den Auswahlen
        filtered_df = aggregated_df.copy()

        # Filter Autor
        if autor != "Alle":
            filtered_df = filtered_df[filtered_df['Autor'] == autor]

        # Filter Titel
        if titel != "Alle" and titel != "Keine Titel verfügbar":
            filtered_df = filtered_df[filtered_df['Titel'] == titel]

        # Filter Jahr
        if jahr != "Alle" and jahr != "Keine Jahre verfügbar":
            try:
                jahr_int = int(jahr)
                filtered_df = filtered_df[filtered_df['Jahr'] == jahr_int]
            except ValueError:
                st.error("Ungültiges Jahr ausgewählt.")
                filtered_df = pd.DataFrame()

        # Filter Monat
        if monat != "Alle" and monat != "Keine Monate verfügbar":
            filtered_df = filtered_df[filtered_df['Monat'] == monat]

        # Filter Währung
        if währung != "Alle" and währung != "Keine Währung verfügbar":
            filtered_df = filtered_df[filtered_df['Währung'] == währung]
        
        # Filter Bonus
        if bonus_filter == "Mit Bonus":
            filtered_df = filtered_df[filtered_df['Bonus'] > 0]
        elif bonus_filter == "Ohne Bonus":
            filtered_df = filtered_df[filtered_df['Bonus'] == 0]
        
        if not filtered_df.empty:
            # Sortiere nach Jahr und Monat_num
            filtered_df = filtered_df.sort_values(by=['Jahr', 'Monat_num'])
            
            # Berechnung der Gesamtmetriken
            total_tantiemen = filtered_df['Tantiemen'].sum()
            total_bonus = filtered_df['Bonus'].sum()
            total_gesamtkäufe = filtered_df['Gesamtverkäufe'].sum()
            total_ebooks = filtered_df['E-Books'].sum()
            total_paperback = filtered_df['Paperback/Hardcover'].sum()
            total_gelesene_seiten = filtered_df['Gelesene Seiten'].sum()
            
            # Mapping der Währungen zu ihren Symbolen
            currency_symbols = {
                'EUR': '€',
                'USD': '$',
                'GBP': '£',
                'CHF': 'CHF',
                'AUD': 'A$',
                'CAD': 'C$',
                'BRL': 'BRL',
                'SEK': 'SEK',
                # Fügen Sie weitere Währungen nach Bedarf hinzu
            }

            # Holen Sie das Symbol basierend auf der ausgewählten Währung
            symbol = currency_symbols.get(währung, '')
    
            # Formatierung der Metriken
            formatted_tantiemen = format_eu_number(total_tantiemen, decimal_places=2) + f" {symbol}"
            formatted_bonus = format_eu_number(total_bonus, decimal_places=2) + f" {symbol}"
            formatted_gesamtkäufe = format_eu_number(total_gesamtkäufe)
            formatted_gelesene_seiten = format_eu_number(total_gelesene_seiten)
            
            # CSS styles für die Metrik-Karten
            st.markdown("""
                <style>
                .stMetric > div {
                    border: 2px solid lightgray;
                    font-size: 25px;
                    padding: 10px;
                    border-radius: 5px;
                    text-align: center;
                }
                </style>
                """, unsafe_allow_html=True)
            
            # Layout für Metriken - Zeile 1
            # st.markdown("### 📈 Metriken Übersicht")
            row1_col1, row1_col2 = st.columns(2)
            with row1_col1:
                st.metric("💰 Gesamteinnahmen", formatted_tantiemen)
            with row1_col2:
                st.metric("🏆 Bonus (in Gesamteinnahmen enthalten)", formatted_bonus)
            
            # Layout für Metriken - Zeile 2
            row2_col1, row2_col2, row2_col3, row2_col4 = st.columns(4)
            with row2_col1:
                st.metric("📊 Gesamtverkäufe", formatted_gesamtkäufe)
            with row2_col2:
                st.metric("📚 E-Books", format_eu_number(total_ebooks))
            with row2_col3:
                st.metric("📖 Paperback/Hardcover", format_eu_number(total_paperback))
            with row2_col4:
                st.metric("📄 Gelesene Seiten", formatted_gelesene_seiten)
                    
            # Zur Gegenkontrolle: Anzeige des gefilterten DataFrames
            st.subheader("📊 Übersicht Verkäufe")
            # Entferne 'Monat_num' aus der Anzeige
            display_df = filtered_df.drop(columns=['Monat_num']).copy()

            # Sicherstellen, dass 'Jahr' integer ist (unabhängig von vorherigen Schritten)
            if 'Jahr' in display_df.columns:
                display_df['Jahr'] = display_df['Jahr'].astype(int)

            # Format 'Tantiemen' und 'Bonus' auf EU-Format mit zwei Dezimalstellen
            if 'Tantiemen' in display_df.columns:
                display_df['Tantiemen'] = display_df['Tantiemen'].apply(lambda x: format_eu_number(x, decimal_places=2))
            if 'Bonus' in display_df.columns:
                display_df['Bonus'] = display_df['Bonus'].apply(lambda x: format_eu_number(x, decimal_places=2))
            
            # Format andere numerische Spalten außer 'Jahr' ohne Dezimalstellen
            for col in ['Gesamtverkäufe', 'E-Books', 'Paperback/Hardcover', 'Gelesene Seiten']:
                if col in display_df.columns:
                    try:
                        display_df[col] = display_df[col].astype(int).apply(lambda x: format_eu_number(x))
                    except:
                        display_df[col] = display_df[col].apply(lambda x: format_eu_number(x))
            
            # Erstellen Sie die Spalte 'Verkaufsmonat'
            display_df['Verkaufsmonat'] = display_df['Monat'] + ' ' + display_df['Jahr'].astype(str)

            # Optional: Entfernen Sie die separaten 'Monat' und 'Jahr' Spalten für eine bessere Darstellung
            display_df = display_df.drop(columns=['Monat', 'Jahr'])
            st.dataframe(display_df, column_order=['Verkaufsmonat', 'Autor', 'Titel', 'Währung', 'Tantiemen', 'Bonus', 'Gesamtverkäufe', 'E-Books', 'Paperback/Hardcover','Gelesene Seiten'], hide_index=True)
            
           # 2. Dynamische Erstellung des Dateinamens beim Download
            if jahr == "Alle":
                if not aggregated_df['Jahr'].empty:
                    year_from = aggregated_df['Jahr'].min()
                    year_to = aggregated_df['Jahr'].max()
                    dateiname = f"{year_from}_bis_{year_to}_Einnahmen"
                else:
                    dateiname = "Einnahmen"
            else:
                if monat != "Alle":
                    dateiname = f"{jahr}_{monat}_Einnahmen"
                else:
                    dateiname = f"{jahr}_Einnahmen"

            # Fügen Sie den Autorennamen hinzu, falls ein spezifischer Autor ausgewählt wurde
            if autor != "Alle" and autor != "Keine Titel verfügbar":
                dateiname += f"_{autor.replace(' ', '_')}"
    
            # Fügen Sie den Titel hinzu, falls ein spezifischer Titel ausgewählt wurde
            if titel != "Alle" and titel != "Keine Titel verfügbar":
                dateiname += f"_{titel.replace(' ', '_')}"
    
            dateiname += ".xlsx"
            
            # Button zum Herunterladen des gefilterten DataFrames als Excel-Datei mit dynamischem Dateinamen
            buffer = io.BytesIO()
            # Inklusive 'Verkaufsmonat' und Ausschluss von 'Monat' & 'Jahr' (bereits entfernt)
            download_df = filtered_df.copy()
            download_df['Verkaufsmonat'] = download_df['Monat'] + ' ' + download_df['Jahr'].astype(str)
            download_df = download_df.drop(columns=['Monat_num', 'Monat', 'Jahr'])
            download_df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            st.download_button(
                label="📥 Download als Excel",
                data=buffer,
                file_name=dateiname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
                st.info("🟡 Keine Daten gefunden für die ausgewählten Filter.")

if __name__ == "__main__":
    main()
